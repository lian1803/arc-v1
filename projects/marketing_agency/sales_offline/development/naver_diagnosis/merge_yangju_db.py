"""
양주 lead_db 9개 파일 통합 — 가장 마지막 (181849) 을 base, 나머지 dedup append.

실행: python merge_yangju_db.py
출력: ../lead_db/양주_통합_{timestamp}.xlsx
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
OUT_PATH = LEAD_DB / f"양주_통합_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def list_yangju_files() -> list[Path]:
    """양주 파일만 (남양주 제외)."""
    files = []
    for p in LEAD_DB.glob("양주_*.xlsx"):
        if p.name.startswith("~$"):
            continue
        if "남양주" in p.name:
            continue
        files.append(p)
    return sorted(files)


def read_rows(path: Path) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    return headers, rows[1:]


def dedup_key(row: dict) -> str:
    """phone 우선, 없으면 name+addr prefix."""
    name = (row.get("업체명") or row.get("상호") or "").strip()
    phone = (row.get("010번호") or row.get("전화") or "").strip()
    addr = (row.get("주소") or "").strip()
    name_norm = re.sub(r"[\s\-_·•]", "", name).lower()
    if phone:
        phone_digits = re.sub(r"\D", "", phone)
        return f"p:{name_norm}:{phone_digits}"
    if addr:
        return f"a:{name_norm}:{addr[:15]}"
    return f"n:{name_norm}"


def main():
    files = list_yangju_files()
    if not files:
        print("양주 파일을 찾을 수 없습니다.")
        return

    # 가장 마지막 timestamp = base. 수집중.xlsx 는 timestamp 없으니 가장 뒤.
    final_files = [f for f in files if "최종" in f.name]
    incomplete = [f for f in files if "수집중" in f.name]
    final_files.sort()  # timestamp 오름차순
    base = final_files[-1] if final_files else files[-1]
    others = [f for f in final_files if f != base] + incomplete

    print(f"[base] {base.name}")
    print(f"[append] {len(others)}개 파일")
    for o in others:
        print(f"  - {o.name}")

    base_headers, base_rows = read_rows(base)
    print(f"\nbase 행수: {len(base_rows)}")
    seen_keys = set()
    merged = []

    # base 먼저
    for row in base_rows:
        if not row or all(c is None for c in row):
            continue
        d = {h: v for h, v in zip(base_headers, row)}
        k = dedup_key(d)
        if k in seen_keys:
            continue
        seen_keys.add(k)
        merged.append(d)
    print(f"base unique: {len(merged)}")

    # 나머지에서 신규 row 만 append
    added_per_file = {}
    for f in others:
        h, rows = read_rows(f)
        before = len(merged)
        for row in rows:
            if not row or all(c is None for c in row):
                continue
            d = {hh: v for hh, v in zip(h, row)}
            k = dedup_key(d)
            if k in seen_keys:
                continue
            seen_keys.add(k)
            merged.append(d)
        added_per_file[f.name] = len(merged) - before
        print(f"  + {f.name}: {added_per_file[f.name]}건 신규")

    # 통합 파일 저장 — base header 사용
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "양주_통합"
    out_ws.append(base_headers)
    for d in merged:
        out_ws.append([d.get(h) for h in base_headers])
    out_wb.save(str(OUT_PATH))

    print(f"\n=== 통합 완료 ===")
    print(f"총 unique: {len(merged)}건")
    print(f"저장: {OUT_PATH}")

    # 간단 통계
    phone_yes = sum(1 for d in merged if d.get("010번호"))
    print(f"010 번호 보유: {phone_yes}건 ({phone_yes*100//max(len(merged),1)}%)")


if __name__ == "__main__":
    main()
