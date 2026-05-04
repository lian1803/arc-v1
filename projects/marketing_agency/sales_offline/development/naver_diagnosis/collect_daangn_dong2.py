"""당근 동별 v2 — 1차 미포함 동 + 2가지 쿼리로 ?in= URL 발굴."""
import asyncio, re, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010[-\s.]?\d{3,4}[-\s.]?\d{4}')
CLEAN = re.compile(r'[^\d]')
DAANGN_RE = re.compile(r'https?://[^\s"\'<>\]]+daangn\.com/kr/local-profile/[^\s"\'<>\]]+')
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"

DONG_MAP = {
    "남양주": ["진건","조안","수동","퇴계원","금곡","양정","평내","지금","도농","삼패"],
    "양주": ["남면","장흥","백석","만송","고읍"],
    "포천": ["영중","창수","내촌","이동","일동","신동","영북","관인"],
    "의정부": ["의정부1동","의정부2동","자금","금오","산곡","민락","용현"],
    "동두천": ["상봉암","탑동","걸산"],
    "가평": ["상면","북면","하면","설악"],
}

QUERIES = [
    "{dong} 당근마켓 local-profile",
    "{dong} 당근 동네업체 전화",
]


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    return f"{d[:3]}-{d[3:7]}-{d[7:]}" if len(d) == 11 else raw


async def find_daangn_urls(ctx, dong: str) -> list[str]:
    url_set = []
    for q_tmpl in QUERIES:
        q = q_tmpl.format(dong=dong)
        pg = await ctx.new_page()
        try:
            await pg.goto(f"https://search.naver.com/search.naver?query={q}",
                          wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(1.0)
            html = await pg.content()
            for m in DAANGN_RE.finditer(html):
                url = m.group(0).rstrip(".,)")
                if url not in url_set:
                    url_set.append(url)
        except PWTimeout:
            pass
        except Exception as e:
            print(f"  [err find] {dong}: {e}")
        finally:
            await pg.close()
        await asyncio.sleep(0.4)
    return url_set


async def scrape_phones(ctx, url: str, region: str, dong: str) -> list[dict]:
    pg = await ctx.new_page()
    results, seen_local = [], set()
    try:
        await pg.goto(url, wait_until="domcontentloaded", timeout=18000)
        await asyncio.sleep(2.5)
        html = await pg.content()
        for m in PHONE_RE.finditer(html):
            d = CLEAN.sub("", m.group(0))
            if len(d) == 11 and d not in seen_local:
                seen_local.add(d)
                results.append({"phone": norm(m.group(0)), "region": region,
                                 "keyword": dong, "name": ""})
    except PWTimeout:
        pass
    except Exception as e:
        print(f"  [err scrape] {dong}: {e}")
    finally:
        await pg.close()
    return results


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in DONG_MAP}

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR",
                                    viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for region, dongs in DONG_MAP.items():
            for dong in dongs:
                urls = await find_daangn_urls(ctx, dong)
                for url in urls[:4]:
                    recs = await scrape_phones(ctx, url, region, dong)
                    await asyncio.sleep(0.8)
                    new = 0
                    for r in recs:
                        d = CLEAN.sub("", r["phone"])
                        if d not in seen:
                            seen.add(d); all_r.append(r); rc[region] += 1; new += 1
                    if new:
                        print(f"  {region}/{dong} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_당근동별2_{TS}.xlsx",
                DESKTOP / f"북부수도권_당근동별2_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg, cnt in rc.items():
        print(f"  {reg}: {cnt}건")


if __name__ == "__main__":
    asyncio.run(main())
