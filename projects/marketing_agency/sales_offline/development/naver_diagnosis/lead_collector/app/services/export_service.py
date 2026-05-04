"""엑셀 내보내기 서비스 — 기존 main_v2.py save_excel() 로직 재활용 + 확장"""
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import EXPORT_DIR


def _resolve_save_path(save_path: str) -> str:
    """저장 경로 결정 — 파일이 열려있어 쓰기 불가면 _1, _2 … 형식으로 대체 경로 반환"""
    p = Path(save_path)
    candidate = p
    suffix = 1
    while True:
        try:
            # 쓰기 가능한지 테스트 (열린 파일은 PermissionError)
            if candidate.exists():
                candidate.open("a").close()
            return str(candidate)
        except PermissionError:
            candidate = p.with_stem(f"{p.stem}_{suffix}")
            suffix += 1
            if suffix > 99:  # 무한 루프 방지
                raise


def to_excel(businesses: list[dict], save_path: str = None) -> str:
    """
    businesses: Business.to_dict() 리스트
    반환: 저장된 xlsx 파일 경로
    """
    if not save_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        region = businesses[0].get("raw_address", "지역")[:10] if businesses else "결과"
        region_clean = re.sub(r'[\\/:*?"<>|]', "_", region)
        save_path = str(EXPORT_DIR / f"수집결과_{region_clean}_{ts}.xlsx")

    save_path = _resolve_save_path(save_path)

    wb = openpyxl.Workbook()

    # 시트 분리: 전체 / 확인됨 / 미확인 / 폐업의심
    sheets = {
        "전체": businesses,
        "확인됨": [b for b in businesses if b.get("verify_status") == "확인됨"],
        "미확인": [b for b in businesses if b.get("verify_status") == "미확인"],
        "폐업의심": [b for b in businesses if b.get("verify_status") == "폐업의심"],
    }

    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        _write_sheet(ws, rows)

    Path(save_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    return save_path


def _write_sheet(ws, rows: list[dict]):
    HEADERS = [
        "업체명", "010번호", "번호상태", "네이버플레이스URL",
        "인스타URL", "당근마켓URL", "수집출처", "업종",
        "주소", "검증상태", "수집일자",
    ]
    HEADER_FILL = PatternFill("solid", fgColor="1a365d")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")

    COLOR_MAP = {
        "확인됨":   "c6efce",
        "미확인":   "ffeb9c",
        "폐업의심": "ffc7ce",
    }
    URL_COLS = {4, 5, 6}  # 1-based

    ws.row_dimensions[1].height = 28
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row_idx, biz in enumerate(rows, 2):
        color = COLOR_MAP.get(biz.get("verify_status", "미확인"), "ffffff")
        fill = PatternFill("solid", fgColor=color)

        row_data = [
            biz.get("name", ""),
            biz.get("phone", ""),
            biz.get("phone_status", ""),
            biz.get("naver_place_url", ""),
            biz.get("insta_url", ""),
            biz.get("daangn_url", ""),
            biz.get("sources", ""),
            biz.get("category", ""),
            biz.get("raw_address", ""),
            biz.get("verify_status", ""),
            biz.get("collected_at", ""),
        ]

        ws.row_dimensions[row_idx].height = 22
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val or "")
            cell.fill = fill
            cell.alignment = LEFT if col in {1, 4, 5, 6, 7, 9} else CENTER
            if col in URL_COLS and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [20, 16, 10, 48, 40, 42, 20, 12, 30, 10, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
