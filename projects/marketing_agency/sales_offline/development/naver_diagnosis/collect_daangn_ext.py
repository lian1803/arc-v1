"""당근 local-profile 확장 — 기존 10 키워드 외 피부관리/세탁소/노래방."""
import asyncio, re, sys, json
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

CLEAN = re.compile(r'[^\d]')
DAANGN_URL = re.compile(r'https?://(?:www\.)?daangn\.com/kr/local-profile/[^\s"\'<>&]+')
REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["피부관리", "세탁소", "노래방"]
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건 로드: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    return raw


def is_mobile(p: str) -> bool:
    d = CLEAN.sub("", p)
    return d.startswith("010") and len(d) == 11


async def get_daangn_urls(page, region: str, kw: str) -> list[str]:
    q = f"{region} {kw} 당근마켓"
    url = f"https://search.naver.com/search.naver?query={q}&where=web"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=18000)
        await page.wait_for_timeout(700)
        html = await page.content()
        return list(dict.fromkeys(DAANGN_URL.findall(html)))
    except Exception:
        return []


async def scrape_profile(page, url: str, region: str, kw: str) -> list[dict]:
    try:
        await page.goto(url, wait_until="load", timeout=15000)
        await page.wait_for_timeout(2500)
        html = await page.content()
        phones = []
        for pat in [r'tel:(01[016789][0-9\-\.]{8,12})',
                    r'010[-\s.]?\d{3,4}[-\s.]?\d{4}']:
            for m in re.finditer(pat, html):
                raw = m.group(1) if m.lastindex else m.group(0)
                p = norm(raw)
                if is_mobile(p) and p not in [x["phone"] for x in phones]:
                    phones.append({"phone": p})
        name_m = re.search(r'"name"\s*:\s*"([^"]{2,40})"', html)
        name = name_m.group(1) if name_m else ""
        if not name:
            h1 = re.search(r'<h1[^>]*>([^<]{2,30})</h1>', html)
            name = h1.group(1) if h1 else ""
        return [{"phone": r["phone"], "name": name, "region": region,
                 "keyword": kw, "daangn_url": url} for r in phones]
    except PWTimeout:
        return []
    except Exception as e:
        print(f"  [err] {url[:50]}: {e}")
        return []


async def main():
    existing = load_existing()
    all_r, seen_phones, seen_urls = [], set(existing), set()
    rc = {r: 0 for r in REGIONS}
    total_q = len(REGIONS) * len(KEYWORDS)
    print(f"\n당근 확장 수집 ({total_q} 쿼리)\n")

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(
            user_agent=UA, locale="ko-KR", viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())
        pg = await ctx.new_page()

        for i, region in enumerate(REGIONS):
            for j, kw in enumerate(KEYWORDS):
                done = i * len(KEYWORDS) + j + 1
                urls = await get_daangn_urls(pg, region, kw)
                new_urls = [u for u in urls if u not in seen_urls][:5]
                for u in new_urls: seen_urls.add(u)
                if new_urls:
                    print(f"  [{done}/{total_q}] {region} {kw} → URL {len(new_urls)}개")
                new_count = 0
                for url in new_urls:
                    recs = await scrape_profile(pg, url, region, kw)
                    await asyncio.sleep(0.6)
                    for r in recs:
                        if r["phone"] not in seen_phones:
                            seen_phones.add(r["phone"])
                            all_r.append(r); rc[region] += 1; new_count += 1
                if new_count:
                    print(f"    +{new_count}건 → 누적 {len(all_r)}건")
                await asyncio.sleep(0.5)

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호", "당근URL"]
    for out in [LEAD_DB / f"북부수도권_당근확장_{TS}.xlsx",
                DESKTOP / f"북부수도권_당근확장_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"], r["daangn_url"]])
        wb.save(str(out))

    bk = LEAD_DB / f"북부수도권_당근확장_{TS}.json"
    with open(str(bk), "w", encoding="utf-8") as f:
        json.dump(all_r, f, ensure_ascii=False, indent=2)

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg in REGIONS: print(f"  {reg}: {rc[reg]}건")


if __name__ == "__main__":
    asyncio.run(main())
