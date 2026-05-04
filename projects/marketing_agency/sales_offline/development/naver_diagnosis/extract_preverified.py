"""통합 file 의 ✅+URL 박힌 row 중 6 region 주소 매칭만 추출 (API X)."""
import sys, glob
from datetime import datetime
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
INPUT = max(LEAD_DB.glob("북부수도권_통합_*.xlsx"), key=lambda p: p.stat().st_mtime)
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
TARGETS = [("남양주", "남양주시"), ("양주", "양주시"), ("포천", "포천시"),
           ("의정부", "의정부시"), ("동두천", "동두천시"), ("가평", "가평군")]


def find_col(headers, *keys):
    for i, h in enumerate(headers):
        if any(k in str(h) for k in keys): return i
    return -1


def matched_region(addr):
    if not addr: return None
    s = str(addr)
    for r, sub in TARGETS:
        if sub in s:
            if r == "양주" and "남양주시" in s: continue
            return r
    return None


def main():
    wb = openpyxl.load_workbook(str(INPUT), read_only=True, data_only=True)
    rows_all = list(wb.active.iter_rows(values_only=True)); wb.close()
    headers = [str(h).strip() if h else "" for h in rows_all[0]]
    rows = rows_all[1:]
    ni = find_col(headers, "업체명")
    kai = find_col(headers, "수집주소")
    nai = find_col(headers, "네이버지도주소")
    npi = find_col(headers, "네이버플레이스")
    vi = find_col(headers, "검증")
    out_h = list(headers) + ["매칭주소", "매칭URL", "매칭지역", "검증상태"]
    verified, unverified = [], []
    by_r = {r: 0 for r, _ in TARGETS}

    print(f"INPUT: {INPUT.name}\n전체: {len(rows)}건")

    for row in rows:
        kaddr = str(row[kai]).strip() if kai >= 0 and row[kai] else ""
        naddr = str(row[nai]).strip() if nai >= 0 and row[nai] else ""
        nurl = str(row[npi]).strip() if npi >= 0 and row[npi] else ""
        vmark = str(row[vi]).strip() if vi >= 0 and row[vi] else ""
        base = list(row)

        # Path 1: ✅ + URL + 6 region 매칭 (naddr OR kaddr)
        if "✅" in vmark and nurl and nurl.startswith("http"):
            tr = matched_region(naddr) or matched_region(kaddr)
            if tr:
                by_r[tr] += 1
                verified.append(base + [naddr or kaddr, nurl, tr, "preverified"])
                continue

        # Path 2: URL 만 박혀있고 (검증 mark 다른) 주소 6 region (둘 다 시도)
        if nurl and nurl.startswith("http"):
            tr = matched_region(naddr) or matched_region(kaddr)
            if tr:
                by_r[tr] += 1
                verified.append(base + [naddr or kaddr, nurl, tr, "url-verified"])
                continue

        # Path 3: kakao 주소만이라도 6 region 매칭 (URL 없음)
        tr = matched_region(kaddr)
        if tr:
            unverified.append(base + [kaddr, "", tr, "addr-only"])
        else:
            unverified.append(base + ["", "", "", "no-match"])

    out_v = LEAD_DB / f"북부수도권_검증v3_{TS}.xlsx"
    out_u = LEAD_DB / f"북부수도권_unverifiedv3_{TS}.xlsx"
    for path, data in [(out_v, verified), (out_u, unverified)]:
        w = openpyxl.Workbook(); s = w.active; s.append(out_h)
        for r in data: s.append(r)
        w.save(str(path))

    print(f"\n=== 결과 ===")
    print(f"verified (URL 박힘 + 6 region): {len(verified)}건")
    print(f"unverified: {len(unverified)}건")
    print(f"지역별 verified:")
    for r, _ in TARGETS:
        print(f"  {r}: {by_r[r]}건")
    print(f"\n저장:\n  {out_v.name}\n  {out_u.name}")


if __name__ == "__main__":
    main()
