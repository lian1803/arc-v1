"""당근마켓 직접 검색 → local_profile 탭 010 추출. 6r × 8kw."""
import asyncio, re, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010[-\s.]?\d{3,4}[-\s.]?\d{4}')
CLEAN = re.compile(r'[^\d]')
IN_RE = re.compile(r'https?://[^\s"\'<>\]]+daangn\.com/kr/local-profile/[^\s"\'<>\]]+\?in=[^\s"\'<>\]&]+')
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"

REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["미용실", "카페", "음식점", "학원", "헬스장", "네일", "피부관리", "고깃집"]


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    return f"{d[:3]}-{d[3:7]}-{d[7:]}" if len(d) == 11 else raw


async def extract_phones(html: str, region: str, kw: str) -> tuple[list[dict], list[str]]:
    found, seen_local, in_urls = [], set(), []
    for m in PHONE_RE.finditer(html):
        d = CLEAN.sub("", m.group(0))
        if len(d) == 11 and d not in seen_local:
            seen_local.add(d)
            found.append({"phone": norm(m.group(0)), "region": region,
                           "keyword": kw, "name": ""})
    for m in IN_RE.finditer(html):
        url = m.group(0).rstrip(".,)")
        if url not in in_urls:
            in_urls.append(url)
    return found, in_urls


async def search_and_collect(ctx, region: str, kw: str,
                              visited: set[str]) -> list[dict]:
    all_found, seen_local = [], set()
    pg = await ctx.new_page()
    try:
        url = f"https://www.daangn.com/search?q={region}+{kw}&tab=local_profile"
        await pg.goto(url, wait_until="domcontentloaded", timeout=18000)
        await asyncio.sleep(3.0)
        html = await pg.content()
        found, in_urls = await extract_phones(html, region, kw)
        for r in found:
            d = CLEAN.sub("", r["phone"])
            if d not in seen_local:
                seen_local.add(d); all_found.append(r)
        for u in in_urls[:3]:
            if u in visited:
                continue
            visited.add(u)
            pg2 = await ctx.new_page()
            try:
                await pg2.goto(u, wait_until="domcontentloaded", timeout=18000)
                await asyncio.sleep(2.5)
                found2, _ = await extract_phones(await pg2.content(), region, kw)
                for r in found2:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen_local:
                        seen_local.add(d); all_found.append(r)
            except Exception:
                pass
            finally:
                await pg2.close()
            await asyncio.sleep(0.8)
    except PWTimeout:
        print(f"  [timeout] {region} {kw}")
    except Exception as e:
        print(f"  [err] {region} {kw}: {e}")
    finally:
        await pg.close()
    return all_found


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in REGIONS}
    visited: set[str] = set()
    total = len(REGIONS) * len(KEYWORDS)
    print(f"\n당근 직접검색 ({total} 쿼리)\n")

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR",
                                    viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for i, region in enumerate(REGIONS):
            for j, kw in enumerate(KEYWORDS):
                done = i * len(KEYWORDS) + j + 1
                recs = await search_and_collect(ctx, region, kw, visited)
                await asyncio.sleep(0.8)
                new = 0
                for r in recs:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen:
                        seen.add(d); all_r.append(r); rc[region] += 1; new += 1
                if new:
                    print(f"  [{done}/{total}] {region} {kw} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_당근직접_{TS}.xlsx",
                DESKTOP / f"북부수도권_당근직접_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg in REGIONS:
        print(f"  {reg}: {rc[reg]}건")


if __name__ == "__main__":
    asyncio.run(main())
