"""Naver 예약(booking.naver.com) → 010 수집. 6 region × 8 keyword."""
import asyncio, re, sys, json
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010\d{7,8}')
CLEAN = re.compile(r'[^\d]')
REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["미용실", "헬스장", "피부관리", "네일", "학원", "카페", "음식점", "마사지"]
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건 로드: {files[-1].name}")
    return seen


def norm(d: str) -> str:
    return f"{d[:3]}-{d[3:7]}-{d[7:]}"


async def search_booking(ctx, region: str, kw: str) -> list[dict]:
    found: list[dict] = []
    seen_d: set[str] = set()
    captured_json: list[str] = []

    pg = await ctx.new_page()
    try:
        async def handle(resp):
            try:
                if resp.status != 200:
                    return
                ct = resp.headers.get("content-type", "")
                if "json" not in ct:
                    return
                body = await resp.json()
                text = json.dumps(body, ensure_ascii=False)
                captured_json.append(text)
            except Exception:
                pass

        pg.on("response", handle)
        url = f"https://booking.naver.com/booking/5/bizes?query={region}+{kw}"
        await pg.goto(url, wait_until="domcontentloaded", timeout=22000)
        await asyncio.sleep(3500 / 1000)
        pg.remove_listener("response", handle)

        # API 응답 JSON 파싱
        for text in captured_json:
            for m in PHONE_RE.finditer(text):
                d = m.group(0)
                if len(d) not in (10, 11) or d in seen_d:
                    continue
                seen_d.add(d)
                ctx_str = text[max(0, m.start() - 200):m.start() + 50]
                nm = re.search(r'"(?:name|bizName|title)"\s*:\s*"([^"]{2,30})"', ctx_str)
                found.append({
                    "phone": norm(d) if len(d) == 11 else f"{d[:3]}-{d[3:6]}-{d[6:]}",
                    "name": nm.group(1) if nm else "",
                    "region": region, "keyword": kw,
                })

        # DOM 폴백
        html = await pg.content()
        for m in PHONE_RE.finditer(html):
            d = m.group(0)
            if len(d) in (10, 11) and d not in seen_d:
                seen_d.add(d)
                found.append({
                    "phone": norm(d) if len(d) == 11 else f"{d[:3]}-{d[3:6]}-{d[6:]}",
                    "name": "", "region": region, "keyword": kw,
                })
    except PWTimeout:
        print(f"  [timeout] {region} {kw}")
    except Exception as e:
        print(f"  [err] {region} {kw}: {e}")
    finally:
        await pg.close()
    return found


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in REGIONS}
    total = len(REGIONS) * len(KEYWORDS)
    print(f"\nNaver 예약 수집 ({total} 쿼리)\n")

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(
            user_agent=UA, locale="ko-KR", viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for i, region in enumerate(REGIONS):
            for j, kw in enumerate(KEYWORDS):
                done = i * len(KEYWORDS) + j + 1
                recs = await search_booking(ctx, region, kw)
                await asyncio.sleep(0.8)
                new = 0
                for r in recs:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen:
                        seen.add(d)
                        all_r.append(r)
                        rc[region] += 1
                        new += 1
                if new:
                    print(f"  [{done}/{total}] {region} {kw} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_네이버예약_{TS}.xlsx",
                DESKTOP / f"북부수도권_네이버예약_{TS}.xlsx"]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    bk = LEAD_DB / f"북부수도권_네이버예약_{TS}.json"
    with open(str(bk), "w", encoding="utf-8") as f:
        json.dump(all_r, f, ensure_ascii=False, indent=2)

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg in REGIONS:
        print(f"  {reg}: {rc[reg]}건")


if __name__ == "__main__":
    asyncio.run(main())
