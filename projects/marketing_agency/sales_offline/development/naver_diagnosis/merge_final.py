"""이번 세션 신규 010 통합 → 업데이트된 FINAL xlsx 생성."""
import sys, re
from datetime import datetime
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

CLEAN = re.compile(r"[^\d]")
DESKTOP = Path.home() / "Desktop"
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# 기존 FINAL 로드
base_files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
if not base_files:
    print("ERR: 010FINAL 파일 없음"); exit(1)
base_file = base_files[-1]
print(f"베이스: {base_file.name}")

wb_base = openpyxl.load_workbook(str(base_file))
ws_base = wb_base.active
hdrs = [c.value for c in ws_base[1]]
existing_rows = list(ws_base.iter_rows(min_row=2, values_only=True))
seen = set()
all_rows = []
for row in existing_rows:
    if row[0]:
        d = CLEAN.sub("", str(row[0]))
        if d not in seen:
            seen.add(d)
            all_rows.append(row)

print(f"기존 {len(all_rows)}건")

# 신규 파일 병합 — 이번 세션 수집본
NEW_PATTERNS = [
    "북부수도권_네이버지역_*.xlsx",
    "북부수도권_네이버맵_*.xlsx",
    "북부수도권_네이버블로그_*.xlsx",
    "북부수도권_당근확장_*.xlsx",
    "북부수도권_인스타그램_*.xlsx",
    "북부수도권_네이버카페_*.xlsx",
    "북부수도권_네이버예약_*.xlsx",
    "북부수도권_당근010_*.xlsx",
    "북부수도권_카카오맵_*.xlsx",
    "북부수도권_구글010_*.xlsx",
    "북부수도권_당근동별_*.xlsx",
    "북부수도권_당근동별2_*.xlsx",
    "북부수도권_당근구글_*.xlsx",
    "북부수도권_네이버지식인_*.xlsx",
    "북부수도권_당근직접_*.xlsx",
    "북부수도권_당근동별3_*.xlsx",
]
new_added = 0
for pat in NEW_PATTERNS:
    for f in sorted(DESKTOP.glob(pat)):
        if f == base_file:
            continue
        wb = openpyxl.load_workbook(str(f))
        ws = wb.active
        file_hdrs = [c.value for c in ws[1]]
        # 010번호 컬럼 위치
        phone_col = next((i for i, h in enumerate(file_hdrs)
                          if h and ("010" in str(h) or "전화" in str(h))), None)
        if phone_col is None:
            phone_col = 0
        name_col = next((i for i, h in enumerate(file_hdrs)
                         if h and "업체" in str(h)), 2)
        region_col = next((i for i, h in enumerate(file_hdrs)
                           if h and "지역" in str(h)), None)
        kw_col = next((i for i, h in enumerate(file_hdrs)
                       if h and "업종" in str(h)), None)

        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            phone = str(row[phone_col]) if row[phone_col] else ""
            d = CLEAN.sub("", phone)
            if not (d.startswith("010") and len(d) == 11):
                continue
            if d in seen:
                continue
            seen.add(d)
            name = str(row[name_col]) if name_col < len(row) and row[name_col] else ""
            region = str(row[region_col]) if region_col is not None and row[region_col] else ""
            kw = str(row[kw_col]) if kw_col is not None and row[kw_col] else ""
            norm = f"{d[:3]}-{d[3:7]}-{d[7:]}"
            # 기존 hdrs: ['010번호', '출처', '지역', '업종', '업체명'] (기존 FINAL 형태 유지)
            src = f.stem.replace("북부수도권_", "").split("_")[0]
            new_row = (norm, src, region, kw, name)
            all_rows.append(new_row)
            cnt += 1
        if cnt:
            print(f"  +{cnt}건 ← {f.name}")
            new_added += cnt

print(f"\n총 {len(all_rows)}건 ({new_added}건 신규)")

# 저장
out_name = f"북부수도권_010FINAL_{TS}.xlsx"
for out_path in [DESKTOP / out_name, LEAD_DB / out_name]:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(hdrs)
    for row in all_rows:
        ws_out.append(list(row))
    wb_out.save(str(out_path))

print(f"저장 완료: {out_name}")
