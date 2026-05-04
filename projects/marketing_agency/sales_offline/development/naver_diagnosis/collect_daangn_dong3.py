"""당근 동별 v3 — v1/v2 미포함 추가 동 + 읍면 단위."""
import asyncio, re, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010[-\s.]?\d{3,4}[-\s.]?\d{4}')
CLEAN = re.compile(r'[^\d]')
DAANGN_RE = re.compile(r'https?://[^\s"\'<>\]]+daangn\.com/kr/local-profile/[^\s"\'<>\]]+')
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"

DONG_MAP = {
    "남양주": ["화도읍","조안면","와부읍","별내면","다산동","오남읍","호평동","마석우동",
               "마석모란동","금남로","퇴계원면"],
    "양주": ["장흥면","은현면","백석읍","광적면","회천동","덕계동","덕정동"],
    "포천": ["포천동","소흘읍","군내면","신북면","가산면","화현면","일동면","이동면"],
    "의정부": ["의정부1동","장암동","신곡1동","신곡2동","호원1동","호원2동","녹양동"],
    "동두천": ["생연1동","생연2동","보산동","중앙동","소요동"],
    "가평": ["가평읍","조종면","청평면","설악면"],
}

QUERIES = [
    "{dong} 당근마켓 local-profile",
    "{dong} 당근마켓 업체",
]


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    return f"{d[:3]}-{d[3:7]}-{d[7:]}" if len(d) == 11 else raw


async def find_daangn_urls(ctx, dong: str) -> list[str]:
    url_set = []
    for q_tmpl in QUERIES:
        pg = await ctx.new_page()
        try:
            q = q_tmpl.format(dong=dong)
            await pg.goto(f"https://search.naver.com/search.naver?query={q}",
                          wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(1.0)
            html = await pg.content()
            for m in DAANGN_RE.finditer(html):
                url = m.group(0).rstrip(".,)")
                if url not in url_set:
                    url_set.append(url)
        except PWTimeout:
            pass
        except Exception as e:
            print(f"  [err find] {dong}: {e}")
        finally:
            await pg.close()
        await asyncio.sleep(0.4)
    return url_set


async def scrape_phones(ctx, url: str, region: str, dong: str) -> list[dict]:
    pg = await ctx.new_page()
    results, seen_local = [], set()
    try:
        await pg.goto(url, wait_until="domcontentloaded", timeout=18000)
        await asyncio.sleep(2.5)
        html = await pg.content()
        for m in PHONE_RE.finditer(html):
            d = CLEAN.sub("", m.group(0))
            if len(d) == 11 and d not in seen_local:
                seen_local.add(d)
                results.append({"phone": norm(m.group(0)), "region": region,
                                 "keyword": dong, "name": ""})
    except PWTimeout:
        pass
    except Exception as e:
        print(f"  [err scrape] {dong}: {e}")
    finally:
        await pg.close()
    return results


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in DONG_MAP}

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR",
                                    viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for region, dongs in DONG_MAP.items():
            for dong in dongs:
                urls = await find_daangn_urls(ctx, dong)
                for url in urls[:4]:
                    recs = await scrape_phones(ctx, url, region, dong)
                    await asyncio.sleep(0.8)
                    new = 0
                    for r in recs:
                        d = CLEAN.sub("", r["phone"])
                        if d not in seen:
                            seen.add(d); all_r.append(r); rc[region] += 1; new += 1
                    if new:
                        print(f"  {region}/{dong} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_당근동별3_{TS}.xlsx",
                DESKTOP / f"북부수도권_당근동별3_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg, cnt in rc.items():
        print(f"  {reg}: {cnt}건")


if __name__ == "__main__":
    asyncio.run(main())
