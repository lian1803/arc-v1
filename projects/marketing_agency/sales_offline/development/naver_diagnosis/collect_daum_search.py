"""Daum 지역검색 → 010 추출. 검증: 양주 5kw × 5업종 = ~30 unique. 6r × 10kw 풀 run."""
import re, sys, requests, urllib3
from datetime import datetime
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
urllib3.disable_warnings()

PHONE_RE = re.compile(r'010[- ]?(\d{4})[- ]?(\d{4})')
NAME_RE = re.compile(r'<a[^>]*class="[^"]*tit_name[^"]*"[^>]*>([^<]+)</a>')
CLEAN = re.compile(r'[^\d]')
REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["미용실", "카페", "음식점", "학원", "헬스장", "네일", "피부관리",
            "고깃집", "치킨", "부동산"]
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
H = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124"}


def load_existing() -> set:
    seen = set()
    for f in LEAD_DB.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        for m in PHONE_RE.finditer(cell):
                            seen.add(f"010{m.group(1)}{m.group(2)}")
            wb.close()
        except Exception:
            pass
    print(f"기존 unique 010: {len(seen):,}건")
    return seen


def search_daum(region: str, kw: str) -> list:
    found, seen_d = [], set()
    for q in [f"{region} {kw} 010", f"{region} {kw} 전화"]:
        try:
            r = requests.get("https://search.daum.net/search",
                             params={"w": "tot", "q": q},
                             timeout=10, headers=H, verify=False)
            if r.status_code != 200:
                continue
            text = r.text
            for m in PHONE_RE.finditer(text):
                d = f"010{m.group(1)}{m.group(2)}"
                if d in seen_d:
                    continue
                seen_d.add(d)
                start = m.start()
                ctx_str = text[max(0, start - 500):start]
                nm_m = NAME_RE.search(ctx_str)
                name = nm_m.group(1).strip() if nm_m else ""
                found.append({"phone": f"010-{m.group(1)}-{m.group(2)}",
                              "name": name, "region": region, "keyword": kw})
        except Exception as e:
            print(f"  [err] {region} {kw}: {type(e).__name__}")
    return found


def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in REGIONS}
    total = len(REGIONS) * len(KEYWORDS)
    print(f"\nDaum 검색 ({total} 쿼리)\n")

    for i, region in enumerate(REGIONS):
        for j, kw in enumerate(KEYWORDS):
            done = i * len(KEYWORDS) + j + 1
            recs = search_daum(region, kw)
            new = 0
            for r in recs:
                d = CLEAN.sub("", r["phone"])
                if d not in seen:
                    seen.add(d)
                    all_r.append(r)
                    rc[region] += 1
                    new += 1
            if new:
                print(f"  [{done}/{total}] {region} {kw} +{new} -> 누적 {len(all_r)}건")

    hdrs = ["지역", "업종", "업체명", "010번호"]
    DESKTOP_DIR = DESKTOP / "오프라인진단"
    DESKTOP_DIR.mkdir(parents=True, exist_ok=True)
    for out in [LEAD_DB / f"북부수도권_다음_{TS}.xlsx",
                DESKTOP_DIR / f"북부수도권_다음_{TS}.xlsx"]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg in REGIONS:
        print(f"  {reg}: {rc[reg]}건")


if __name__ == "__main__":
    main()
