"""구글로 당근 ?in= 집계 페이지 URL 발굴 → 010 추출."""
import asyncio, re, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010[-\s.]?\d{3,4}[-\s.]?\d{4}')
CLEAN = re.compile(r'[^\d]')
DAANGN_IN_RE = re.compile(
    r'https?://[^\s"\'<>\]]+daangn\.com/kr/local-profile/[^\s"\'<>\]]+\?in=[^\s"\'<>\]&]+')
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"

REGIONS = [
    "남양주시", "양주시", "포천시", "의정부시", "동두천시", "가평군",
    "남양주", "양주", "포천", "의정부", "동두천", "가평",
]
RC_BASE = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    return f"{d[:3]}-{d[3:7]}-{d[7:]}" if len(d) == 11 else raw


async def google_find_daangn(ctx, region: str) -> list[str]:
    pg = await ctx.new_page()
    urls = []
    try:
        q = f'site:daangn.com "local-profile" "in=" {region} 업체'
        await pg.goto(f"https://www.google.com/search?q={q}&num=20&hl=ko",
                      wait_until="domcontentloaded", timeout=18000)
        await asyncio.sleep(2.0)
        html = await pg.content()
        if "captcha" in html.lower() or "unusual traffic" in html.lower():
            return []
        for m in DAANGN_IN_RE.finditer(html):
            url = m.group(0).rstrip(".,)")
            if url not in urls:
                urls.append(url)
        for m in re.finditer(
                r'daangn\.com/kr/local-profile/[^\s"\'<>&\]]+\?in=[^\s"\'<>&\]]+', html):
            full = "https://www." + m.group(0).rstrip(".,)")
            if full not in urls:
                urls.append(full)
    except PWTimeout:
        pass
    except Exception as e:
        print(f"  [err google] {region}: {e}")
    finally:
        await pg.close()
    return urls


async def scrape_phones(ctx, url: str, region: str) -> list[dict]:
    pg = await ctx.new_page()
    results, seen_local = [], set()
    try:
        await pg.goto(url, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(2.5)
        html = await pg.content()
        for m in PHONE_RE.finditer(html):
            d = CLEAN.sub("", m.group(0))
            if len(d) == 11 and d not in seen_local:
                seen_local.add(d)
                results.append({"phone": norm(m.group(0)), "region": region,
                                 "keyword": "당근동네", "name": ""})
    except PWTimeout:
        pass
    except Exception as e:
        print(f"  [err scrape]: {e}")
    finally:
        await pg.close()
    return results


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in RC_BASE}
    visited: set[str] = set()

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR",
                                    viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for region in REGIONS:
            base = region.replace("시", "").replace("군", "")
            urls = await google_find_daangn(ctx, region)
            await asyncio.sleep(1.5)
            for url in urls[:6]:
                if url in visited:
                    continue
                visited.add(url)
                recs = await scrape_phones(ctx, url, base)
                await asyncio.sleep(1.0)
                new = 0
                for r in recs:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen:
                        seen.add(d); all_r.append(r)
                        if base in rc:
                            rc[base] += 1
                        new += 1
                if new:
                    print(f"  {base} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_당근구글_{TS}.xlsx",
                DESKTOP / f"북부수도권_당근구글_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg, cnt in rc.items():
        print(f"  {reg}: {cnt}건")


if __name__ == "__main__":
    asyncio.run(main())
