"""
CRM 서비스 — sales_crm/app.py 헬퍼 함수 포팅
"""
from __future__ import annotations

import os
import re
import json
import sqlite3
import asyncio
import importlib.util as _ilu
from pathlib import Path
from datetime import datetime, date
from typing import Optional

from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent.parent  # naver-diagnosis/
NAVER_DIAGNOSIS_DB = BASE_DIR / "diagnosis.db"
EXCEL_DB_FOLDER = BASE_DIR.parent.parent / "DB"
DESKTOP = os.path.expanduser("~/Desktop")
SALES_DIR = os.path.join(DESKTOP, "영업")

# .env 로드
for _p in BASE_DIR.resolve().parents:
    if (_p / "company" / ".env").exists():
        load_dotenv(_p / "company" / ".env")
        break
else:
    load_dotenv()

import google.generativeai as genai
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# message_generator 로드 (Playwright 의존성 회피)
try:
    _mg_spec = _ilu.spec_from_file_location(
        "msg_gen_crm", str(BASE_DIR / "services" / "message_generator.py")
    )
    _mg_mod = _ilu.module_from_spec(_mg_spec)
    _mg_spec.loader.exec_module(_mg_mod)
    _msg_gen_first = _mg_mod.generate_first_message
    _msg_gen_fourth = _mg_mod.generate_fourth_messages
    MESSAGE_GENERATOR_AVAILABLE = True
except Exception as _e:
    print(f"[CRM] message_generator 로드 실패: {_e}")
    MESSAGE_GENERATOR_AVAILABLE = False

# ──────────────────────────────────────────────────
# 카테고리 추정
# ──────────────────────────────────────────────────

def infer_category_from_name(name: str) -> str:
    keywords = {
        "미용실": ["헤어", "미용", "매직", "파마", "스타일"],
        "네일": ["네일", "손톱", "페디"],
        "식당": ["음식", "밥", "국수", "카레", "숯불"],
        "카페": ["카페", "커피", "아메리카노", "라떼"],
        "피부관리": ["피부", "관리", "에스테틱", "스킨", "여드름"],
        "학원": ["학원", "영어", "수학", "과외", "튜터"],
    }
    name_lower = name.lower()
    for category, words in keywords.items():
        if any(word in name_lower for word in words):
            return category
    return "기타"


def get_default_diagnosis_data(name: str, phone: str) -> dict:
    return {
        "business_name": name,
        "category": infer_category_from_name(name),
        "review_count": 0,
        "photo_count": 0,
        "keyword_score": 30,
        "review_score": 20,
        "blog_score": 10,
        "info_score": 25,
        "photo_score": 15,
        "convenience_score": 25,
        "engagement_score": 10,
        "total_score": 18,
        "grade": "D",
        "naver_place_rank": 15,
        "competitor_avg_review": 0,
        "competitor_avg_photo": 0,
        "estimated_lost_customers": 0,
        "has_kakao": False,
        "has_instagram": False,
        "has_hours": False,
        "has_menu": False,
        "has_price": False,
        "has_intro": False,
        "has_directions": False,
        "has_owner_reply": False,
        "news_last_days": 90,
        "keywords": [],
    }

# ──────────────────────────────────────────────────
# 주소 → 지역 추출
# ──────────────────────────────────────────────────

def extract_region_from_address(address: str) -> str:
    if not address:
        return ""
    parts = address.split()
    if len(parts) >= 2:
        region = parts[1]
        region = re.sub(r'(시|군|구)$', '', region)
        return region
    return ""

# ──────────────────────────────────────────────────
# Excel 전화번호 캐시
# ──────────────────────────────────────────────────

_PHONE_LOOKUP_CACHE: dict = {}


def build_phone_lookup() -> dict:
    global _PHONE_LOOKUP_CACHE
    if _PHONE_LOOKUP_CACHE:
        return _PHONE_LOOKUP_CACHE

    def detect_phone_col(headers):
        for i, h in enumerate(headers):
            if h in ('010번호', '전화번호'):
                return i
        for i, h in enumerate(headers):
            if '010' in h:
                return i
        for i, h in enumerate(headers):
            if '전화' in h:
                return i
        return None

    def detect_name_col(headers):
        for i, h in enumerate(headers):
            if h in ('업체명', '상호명', '업체이름', '이름'):
                return i
        for i, h in enumerate(headers):
            if '업체' in h or '상호' in h:
                return i
        return None

    try:
        import openpyxl
        import glob as _glob
        escaped_base = _glob.escape(str(EXCEL_DB_FOLDER))
        xlsx_files = (
            list(_glob.glob(escaped_base + "/db/*.xlsx")) +
            list(_glob.glob(escaped_base + "/output/*.xlsx")) +
            list(_glob.glob(escaped_base + "/*.xlsx"))
        )

        lookup = {}
        for f in xlsx_files:
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                raw_headers = [c.value for c in next(ws.iter_rows())]
                headers = [str(h) if h else '' for h in raw_headers]

                name_col = detect_name_col(headers)
                phone_col = detect_phone_col(headers)
                addr_col = next((i for i, h in enumerate(headers) if '주소' in h), None)

                if name_col is None or phone_col is None:
                    wb.close()
                    continue

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= max(name_col, phone_col):
                        continue
                    name = str(row[name_col]).strip() if row[name_col] else ''
                    phone = str(row[phone_col]).strip() if row[phone_col] else ''
                    addr = str(row[addr_col]).strip() if addr_col is not None and row[addr_col] else ''
                    if name and phone and phone.startswith('010') and name not in lookup:
                        lookup[name] = (phone, addr)
                wb.close()
            except Exception:
                pass

        _PHONE_LOOKUP_CACHE = lookup
        print(f"[CRM] Excel 전화번호 캐시 구축: {len(lookup)}개 업체")
        return lookup
    except Exception as e:
        print(f"[CRM] Excel 전화번호 캐시 구축 실패: {e}")
        return {}


def get_phone_from_excel(name: str):
    lookup = build_phone_lookup()
    if name in lookup:
        return lookup[name]
    name_stripped = name.replace(' ', '')
    for key, val in lookup.items():
        if key.replace(' ', '') == name_stripped:
            return val
    return None, None

# ──────────────────────────────────────────────────
# 파일명 정규화
# ──────────────────────────────────────────────────

def normalize_filename(name: str, phone: str, region: str = "") -> str:
    safe_name = re.sub(r'[\\/:*?"<>|]', '', name).strip() or "업체"
    safe_phone = ''
    if phone and not phone.startswith('진단') and phone != 'unknown':
        safe_phone = phone
    if not region:
        region = os.getenv("SALES_REGION", "")
    safe_region = re.sub(r'[\\/:*?"<>|]', '', region).strip() if region else ""
    parts = []
    if safe_region:
        parts.append(safe_region)
    parts.append(safe_name)
    if safe_phone:
        parts.append(safe_phone)
    return '-'.join(parts)


def extract_business_info(filename: str):
    name = Path(filename).stem
    parts = name.split('_')
    if not parts:
        return None, None
    phone = parts[0]
    business_name = '_'.join(parts[1:]) if len(parts) > 1 else '미등록'
    return phone, business_name


def _get_business_folder(name: str, phone: str, region: str):
    folder_name = normalize_filename(name, phone, region)
    folder_path = os.path.join(SALES_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path, folder_name

# ──────────────────────────────────────────────────
# naver-diagnosis DB 조회 (sync)
# ──────────────────────────────────────────────────

def _fetch_diagnosis_row(c, name: str):
    """4단계 매칭 전략으로 diagnosis_history 행 조회"""
    regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']

    for with_competitor in [True, False]:
        cond = "AND competitor_avg_review > 0" if with_competitor else ""

        # 1. 정확 매칭
        c.execute(f"SELECT * FROM diagnosis_history WHERE business_name = ? {cond} ORDER BY created_at DESC LIMIT 1", (name,))
        row = c.fetchone()
        if row:
            return row

        # 2. 공백 제거
        c.execute(f"SELECT * FROM diagnosis_history WHERE REPLACE(business_name,' ','') = REPLACE(?,' ','') {cond} ORDER BY created_at DESC LIMIT 1", (name,))
        row = c.fetchone()
        if row:
            return row

        # 3. 부분 매칭
        c.execute(f"SELECT * FROM diagnosis_history WHERE business_name LIKE ? {cond} ORDER BY created_at DESC LIMIT 1", (f"%{name}%",))
        row = c.fetchone()
        if row:
            return row

        # 4. 지역명 제거
        for region in regions:
            cleaned = name.replace(region, '').strip()
            if cleaned and cleaned != name:
                c.execute(f"SELECT * FROM diagnosis_history WHERE business_name LIKE ? {cond} ORDER BY created_at DESC LIMIT 1", (f"%{cleaned}%",))
                row = c.fetchone()
                if row:
                    return row
    return None


def get_diagnosis_data_sync(name: str) -> Optional[dict]:
    try:
        if not NAVER_DIAGNOSIS_DB.exists():
            return None
        conn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        row = _fetch_diagnosis_row(c, name)
        conn.close()
        if not row:
            return None
        data = dict(row)
        for field in ('keywords', 'improvement_points'):
            if isinstance(data.get(field), str):
                try:
                    data[field] = json.loads(data[field])
                except Exception:
                    data[field] = []
        return data
    except Exception as e:
        print(f"[CRM] diagnosis DB 조회 실패: {e}")
        return None


def get_naver_diagnosis_by_name_sync(name: str) -> Optional[dict]:
    """id, place_url, ppt_filename, total_score, grade 반환"""
    try:
        if not NAVER_DIAGNOSIS_DB.exists():
            return None
        conn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']
        cols = "id, place_url, ppt_filename, total_score, grade"

        for with_competitor in [True, False]:
            cond = "AND competitor_avg_review > 0" if with_competitor else ""
            for sql, params in [
                (f"SELECT {cols} FROM diagnosis_history WHERE business_name = ? {cond} ORDER BY created_at DESC LIMIT 1", (name,)),
                (f"SELECT {cols} FROM diagnosis_history WHERE REPLACE(business_name,' ','')=REPLACE(?,' ','') {cond} ORDER BY created_at DESC LIMIT 1", (name,)),
                (f"SELECT {cols} FROM diagnosis_history WHERE business_name LIKE ? {cond} ORDER BY created_at DESC LIMIT 1", (f"%{name}%",)),
            ]:
                c.execute(sql, params)
                row = c.fetchone()
                if row:
                    conn.close()
                    return dict(row)

            for region in regions:
                cleaned = name.replace(region, '').strip()
                if cleaned and cleaned != name:
                    c.execute(f"SELECT {cols} FROM diagnosis_history WHERE business_name LIKE ? {cond} ORDER BY created_at DESC LIMIT 1", (f"%{cleaned}%",))
                    row = c.fetchone()
                    if row:
                        conn.close()
                        return dict(row)

        conn.close()
        return None
    except Exception as e:
        print(f"[CRM] naver_diagnosis 조회 실패: {e}")
        return None

# ──────────────────────────────────────────────────
# 거절 유형 감지
# ──────────────────────────────────────────────────

def detect_rejection_type(message: str) -> str:
    msg_lower = message.lower()
    if any(w in msg_lower for w in ["비싸", "비용", "금액", "가격", "비쌈", "너무", "많아"]):
        return "비싸다"
    if any(w in msg_lower for w in ["직접", "혼자", "제가 할", "우리가", "할 수", "스스로"]):
        return "직접"
    if any(w in msg_lower for w in ["나중에", "생각해", "괜찮아", "됐어", "안 돼", "아직"]):
        return "보류"
    if any(w in msg_lower for w in ["해봤", "효과", "없었", "전에", "이미", "안 됐"]):
        return "경험있음"
    return "기타"

# ──────────────────────────────────────────────────
# AI 메시지 생성
# ──────────────────────────────────────────────────

def _clean_markdown(text: str) -> str:
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'\*\*', '', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    return text.strip()


def generate_crm_first_message(name: str, phone: str) -> str:
    try:
        diagnosis_data = get_diagnosis_data_sync(name)
        if not diagnosis_data:
            diagnosis_data = get_default_diagnosis_data(name, phone)

        # 캐시 확인
        if diagnosis_data.get('ai_first_message'):
            return diagnosis_data['ai_first_message']

        # 카테고리 보정
        if not diagnosis_data.get('category') or diagnosis_data.get('category') == '기타':
            diagnosis_data['category'] = infer_category_from_name(name)
        if not diagnosis_data.get('naver_place_rank') or diagnosis_data['naver_place_rank'] <= 0:
            diagnosis_data['naver_place_rank'] = 15

        generated = None

        # 진단 데이터 없는 Excel 업체 → Gemini 우선
        is_excel_only = (
            diagnosis_data.get('naver_place_rank', 0) == 15 and
            diagnosis_data.get('review_count', 0) == 0 and
            not diagnosis_data.get('competitor_avg_review', 0)
        )

        if is_excel_only:
            # Gemini로 직접 생성
            try:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 소상공인 사장님께 카카오톡 첫 메시지를 써줘.

업체명: {name}
상황: 네이버 플레이스 관리가 미흡한 상태로 추정

규칙: 마크다운 완전 금지. 5줄 이내. 자기소개 없음. 업체명으로 시작. 끝에 한 가지 질문으로 마무리. 완성본만."""
                model = genai.GenerativeModel('gemini-2.5-flash')
                response = model.generate_content(prompt)
                generated = _clean_markdown(response.text)
                if generated:
                    return generated
            except Exception as e:
                print(f"[CRM] Gemini 1차 생성 실패 (Excel업체): {e}")
            # fallback to rule-based below

        if MESSAGE_GENERATOR_AVAILABLE and not generated:
            try:
                result = _msg_gen_first(diagnosis_data)
                generated = result.get("text", "") if isinstance(result, dict) else str(result)
            except Exception as e:
                print(f"[CRM] message_generator 1차 실패: {e}")

        if not generated:
            rank = diagnosis_data.get('naver_place_rank', 15)
            my_review = diagnosis_data.get('review_count', 0)
            comp_review = diagnosis_data.get('competitor_avg_review', 0)
            category = diagnosis_data.get('category', '업종')
            region = os.environ.get('SALES_REGION', '')
            search_q = f"{region} {category}".strip() if region else category

            if comp_review > 0 and comp_review >= my_review * 2:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: 주변 상위 가게 리뷰 평균 {comp_review}개, {name}은 {my_review}개

규칙: 마크다운 완전 금지. 5줄 이내. 자기소개 없음. "혹시 이 차이 눈에 띄신 적 있으세요?" 로 끝내기. 완성본만."""
            elif rank > 0:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: "{search_q}" 검색 시 {rank}위권

규칙: 마크다운 완전 금지. 5줄 이내. 자기소개 없음. "혹시 이런 부분 신경 쓰고 계신 편인가요?" 로 끝내기. 완성본만."""
            else:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: 네이버 플레이스 관리 상태 미흡

규칙: 마크다운 완전 금지. 5줄 이내. 자기소개 없음. 질문으로 끝내기. 완성본만."""

            model = genai.GenerativeModel('gemini-2.5-flash')
            response = model.generate_content(prompt)
            generated = response.text

        if generated:
            generated = _clean_markdown(generated)

        # 캐시 저장
        if generated and NAVER_DIAGNOSIS_DB.exists():
            try:
                nc = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
                nc.execute(
                    "UPDATE diagnosis_history SET ai_first_message = ? WHERE business_name = ? AND ai_first_message IS NULL",
                    (generated, name)
                )
                nc.commit()
                nc.close()
            except Exception:
                pass

        return generated or f"{name} 사장님, 데이터 정리하다가 연락드렸습니다. 확인 가능하신가요?"
    except Exception as e:
        print(f"[CRM] 1차 메시지 생성 실패: {e}")
        return f"{name} 사장님, 데이터 정리하다가 연락드렸습니다. 확인 가능하신가요?"


def generate_crm_response_messages(name: str, phone: str, customer_message: str) -> str:
    try:
        rejection_type = detect_rejection_type(customer_message)
        diagnosis_data = get_diagnosis_data_sync(name) or get_default_diagnosis_data(name, phone)

        if MESSAGE_GENERATOR_AVAILABLE and rejection_type != "기타":
            try:
                fourth_messages = _msg_gen_fourth(diagnosis_data)
                if rejection_type in fourth_messages and fourth_messages[rejection_type]:
                    return fourth_messages[rejection_type]
            except Exception as e:
                print(f"[CRM] message_generator 4차 실패: {e}")

        prompt = f"""
너는 실전 영업의 대가야. 네이버 플레이스 마케팅 대행사의 베테랑 영업사원.

[업체 정보]
- 업체명: {name}
- 종합점수: {diagnosis_data.get('total_score', 0):.0f}점 ({diagnosis_data.get('grade', 'D')}등급)

[고객 메시지]
{customer_message if customer_message else "(무응답)"}

규칙: 구체적 해결책 금지. 월액만. 자기소개 금지. 3~5줄. 메시지만 출력.
"""
        model = genai.GenerativeModel('gemini-2.5-flash')
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        print(f"[CRM] 응답 메시지 생성 실패: {e}")
        return "그 말씀이 현실적인 얘기네요. 혹시 시간 10분만 내주셔서 계산 한번 확인해보실래요?"

# ──────────────────────────────────────────────────
# CRM DB 동기화 (diagnosis_history → crm_businesses)
# ──────────────────────────────────────────────────

async def sync_from_naver_diagnosis_to_crm(db) -> int:
    """diagnosis_history의 업체를 crm_businesses에 동기화"""
    from sqlalchemy import select
    from models import DiagnosisHistory, CRMBusiness

    try:
        result = await db.execute(select(DiagnosisHistory).order_by(DiagnosisHistory.created_at.desc()))
        histories = result.scalars().all()
        if not histories:
            return 0

        added = 0
        for h in histories:
            if not h.business_name:
                continue

            existing = await db.execute(
                select(CRMBusiness).where(CRMBusiness.name == h.business_name)
            )
            if existing.scalar_one_or_none():
                continue

            phone_val, addr_val = get_phone_from_excel(h.business_name)
            phone_val = phone_val or f"진단-{h.id}"
            addr_val = addr_val or h.address or ""

            # phone unique 충돌 방지
            phone_check = await db.execute(
                select(CRMBusiness).where(CRMBusiness.phone == phone_val)
            )
            if phone_check.scalar_one_or_none():
                phone_val = f"진단-{h.id}"

            biz = CRMBusiness(
                phone=phone_val,
                name=h.business_name,
                location=addr_val,
                place_url=h.place_url,
                naver_diagnosis_id=h.id,
                status="1차_발송_대기",
            )
            db.add(biz)
            added += 1

        await db.commit()
        if added:
            print(f"[CRM] {added}개 업체 동기화 완료")
        return added
    except Exception as e:
        print(f"[CRM] 동기화 실패: {e}")
        await db.rollback()
        return 0

# ──────────────────────────────────────────────────
# PDF 감시 asyncio 태스크
# ──────────────────────────────────────────────────

async def _add_crm_business_async(phone: str, name: str, pdf_filename: str):
    from database import async_session_maker
    from models import CRMBusiness
    from sqlalchemy import select

    async with async_session_maker() as db:
        try:
            existing = await db.execute(select(CRMBusiness).where(CRMBusiness.phone == phone))
            if existing.scalar_one_or_none():
                return

            diag = get_naver_diagnosis_by_name_sync(name)
            place_url = diag.get('place_url') if diag else None
            naver_id = diag.get('id') if diag else None

            biz = CRMBusiness(
                phone=phone,
                name=name,
                pdf_path=pdf_filename,
                place_url=place_url,
                naver_diagnosis_id=naver_id,
                status="1차_발송_대기",
            )
            db.add(biz)
            await db.commit()
            print(f"[PDF Monitor] 업체 추가: {phone} | {name}")
        except Exception as e:
            print(f"[PDF Monitor] 업체 추가 실패: {e}")
            await db.rollback()


async def monitor_pdfs_async():
    """asyncio PDF 폴더 감시 (5초마다)"""
    monitored: set = set()
    while True:
        try:
            if BASE_DIR.exists():
                for pdf_file in BASE_DIR.glob("*.pdf"):
                    if pdf_file.name not in monitored:
                        phone, name = extract_business_info(pdf_file.name)
                        if phone and name:
                            await _add_crm_business_async(phone, name, pdf_file.name)
                            monitored.add(pdf_file.name)
        except Exception as e:
            print(f"[PDF Monitor] 오류: {e}")
        await asyncio.sleep(5)

# ──────────────────────────────────────────────────
# 지역 목록 및 엑셀 DB 연동 (2026-04-20)
# ──────────────────────────────────────────────────

def scan_available_regions() -> list:
    """DB 폴더에서 사용 가능한 지역 목록 추출"""
    try:
        import glob as _glob
        db_folder = Path(__file__).parent.parent.parent.parent / "DB"
        if not db_folder.exists():
            return []

        regions = set()
        pattern = str(db_folder / "*_010번호_최종_*.xlsx")
        for filepath in _glob.glob(pattern):
            filename = Path(filepath).name
            region_name = filename.split('_')[0]
            if region_name and region_name not in regions:
                regions.add(region_name)

        return sorted(list(regions))
    except Exception as e:
        print(f"[CRM] 지역 스캔 실패: {e}")
        return []


def get_latest_excel_for_region(region: str) -> Optional[Path]:
    """지역별 최신 엑셀 파일 반환"""
    try:
        db_folder = Path(__file__).parent.parent.parent.parent / "DB"
        if not db_folder.exists():
            return None

        import glob as _glob
        pattern = str(db_folder / f"{region}_010번호_최종_*.xlsx")
        files = sorted(_glob.glob(pattern))
        if files:
            return Path(files[-1])
        return None
    except Exception as e:
        print(f"[CRM] 엑셀 파일 조회 실패 ({region}): {e}")
        return None


def read_excel_businesses(excel_path: Path) -> list:
    """엑셀에서 업체 목록 읽기 (dict 리스트로 반환)"""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # 헤더 읽기
        raw_headers = [c.value for c in next(ws.iter_rows())]
        headers = [str(h).strip() if h else '' for h in raw_headers]

        # 컬럼 감지
        phone_col = next((i for i, h in enumerate(headers) if h in ('010번호', '전화번호')), None)
        name_col = next((i for i, h in enumerate(headers) if h in ('업체명', '상호명')), None)
        industry_col = next((i for i, h in enumerate(headers) if h in ('업종',)), None)
        addr_col = next((i for i, h in enumerate(headers) if h in ('주소',)), None)

        if phone_col is None or name_col is None:
            print(f"[CRM] 엑셀 헤더 감지 실패: {excel_path.name}")
            wb.close()
            return []

        businesses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(phone_col, name_col):
                continue

            phone = str(row[phone_col]).strip() if row[phone_col] else ''
            name = str(row[name_col]).strip() if row[name_col] else ''
            industry = str(row[industry_col]).strip() if industry_col and row[industry_col] else ''
            addr = str(row[addr_col]).strip() if addr_col and row[addr_col] else ''

            if phone and name and phone.startswith('010'):
                businesses.append({
                    'phone': phone,
                    'name': name,
                    'industry': industry,
                    'address': addr,
                })

        wb.close()
        return businesses
    except Exception as e:
        print(f"[CRM] 엑셀 읽기 실패: {e}")
        return []


async def load_region_businesses_to_db(db, region: str) -> list:
    """지역 엑셀 업체들을 CRM DB에 로드 (중복 없이, bulk 처리)"""
    from sqlalchemy import select
    from models import CRMBusiness

    excel_path = get_latest_excel_for_region(region)
    if not excel_path:
        return []

    excel_data = read_excel_businesses(excel_path)
    if not excel_data:
        return []

    # 이 지역 기존 업체 조회
    region_result = await db.execute(
        select(CRMBusiness).where(CRMBusiness.region == region)
    )
    region_bizs = region_result.scalars().all()
    region_phones = {b.phone for b in region_bizs}
    phone_to_biz = {b.phone: b for b in region_bizs}

    # 전체 DB phone 조회 (타 지역 중복 INSERT 방지)
    all_phones_result = await db.execute(select(CRMBusiness.phone))
    all_existing_phones = {r[0] for r in all_phones_result.all()}

    result_businesses = [phone_to_biz[b.phone].to_dict() for b in region_bizs]

    new_bizs = []
    for biz_data in excel_data:
        phone = biz_data['phone']
        if phone in all_existing_phones:
            # 이미 다른 지역에 있는 번호면 skip (중복 INSERT 방지)
            continue
        crm_biz = CRMBusiness(
            phone=phone,
            name=biz_data['name'],
            excel_phone=phone,
            industry=biz_data['industry'],
            location=biz_data['address'],
            region=region,
            status="1차_발송_대기",
        )
        new_bizs.append(crm_biz)
        all_existing_phones.add(phone)

    if new_bizs:
        db.add_all(new_bizs)
        await db.commit()
        for b in new_bizs:
            await db.refresh(b)
            result_businesses.append(b.to_dict())

    print(f"[CRM] {region} 지역 {len(result_businesses)}개 (신규 {len(new_bizs)}개)")
    return result_businesses
