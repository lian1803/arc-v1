"""네이버 지식iN 검색 → 답변 속 010 추출. 6r × 8kw."""
import asyncio, re, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PHONE_RE = re.compile(r'010[-\s.]?\d{3,4}[-\s.]?\d{4}')
CLEAN = re.compile(r'[^\d]')
REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["미용실", "음식점", "카페", "학원", "헬스장", "네일", "피부관리", "고깃집"]
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    seen = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            seen.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(seen)}건: {files[-1].name}")
    return seen


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    return f"{d[:3]}-{d[3:7]}-{d[7:]}" if len(d) == 11 else raw


async def search_kin(ctx, region: str, kw: str) -> list[dict]:
    results, seen_local = [], set()
    pg = await ctx.new_page()
    try:
        url = (f"https://search.naver.com/search.naver"
               f"?where=kin&query={region}+{kw}+전화번호+추천&display=10")
        await pg.goto(url, wait_until="domcontentloaded", timeout=16000)
        await asyncio.sleep(1.5)
        html = await pg.content()
        for m in PHONE_RE.finditer(html):
            d = CLEAN.sub("", m.group(0))
            if len(d) == 11 and d not in seen_local:
                seen_local.add(d)
                ctx_str = html[max(0, m.start() - 150):m.start()]
                nm = re.search(r'>([가-힣a-zA-Z0-9 ]{2,15})</(?:span|div|a|b)', ctx_str)
                results.append({"phone": norm(m.group(0)),
                                 "name": nm.group(1).strip() if nm else "",
                                 "region": region, "keyword": kw})
    except PWTimeout:
        print(f"  [timeout] {region} {kw}")
    except Exception as e:
        print(f"  [err] {region} {kw}: {e}")
    finally:
        await pg.close()
    return results


async def main():
    existing = load_existing()
    all_r, seen = [], set(existing)
    rc = {r: 0 for r in REGIONS}
    total = len(REGIONS) * len(KEYWORDS)
    print(f"\n네이버 지식iN ({total} 쿼리)\n")

    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR",
                                    viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())

        for i, region in enumerate(REGIONS):
            for j, kw in enumerate(KEYWORDS):
                done = i * len(KEYWORDS) + j + 1
                recs = await search_kin(ctx, region, kw)
                await asyncio.sleep(0.6)
                new = 0
                for r in recs:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen:
                        seen.add(d); all_r.append(r); rc[region] += 1; new += 1
                if new:
                    print(f"  [{done}/{total}] {region} {kw} +{new} → 누적 {len(all_r)}건")

        await br.close()

    hdrs = ["지역", "업종", "업체명", "010번호"]
    for out in [LEAD_DB / f"북부수도권_네이버지식인_{TS}.xlsx",
                DESKTOP / f"북부수도권_네이버지식인_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r:
            ws.append([r["region"], r["keyword"], r["name"], r["phone"]])
        wb.save(str(out))

    print(f"\n=== 완료 ===\n신규 010: {len(all_r)}건")
    for reg in REGIONS:
        print(f"  {reg}: {rc[reg]}건")


if __name__ == "__main__":
    asyncio.run(main())
