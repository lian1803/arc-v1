"""6개 지역 (양주/포천/의정부/남양주/동두천/가평) lead_db 합치기.
각 row 에 _지역 태그 추가 + dedup by (이름+phone)."""
import re, sys
from datetime import datetime
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
OUT = LEAD_DB / f"북부수도권_통합_{TS}.xlsx"
REGIONS = ["양주", "포천", "의정부", "남양주", "동두천", "가평"]


def list_files():
    """timestamp 내림차순 정렬 — 가장 최신 (수집주소 populated) file 이 base 로."""
    by_region = {r: [] for r in REGIONS}
    for r in REGIONS:
        for p in LEAD_DB.glob(f"{r}_*.xlsx"):
            if any(k in p.name for k in ("통합", "검증", "unverified")):
                continue
            if p.name.startswith("~$"):
                continue
            if not p.name.startswith(f"{r}_"):
                continue
            by_region[r].append(p)
    def _key(p):
        # primary: 최종 우선 (0), 수집중 뒤로 (1)
        primary = 1 if "수집중" in p.name else 0
        m = re.search(r'(\d{8}_\d{6})', p.name)
        ts = int((m.group(1) if m else "0").replace('_', ''))
        return (primary, -ts)  # ts 내림차순 (최신 먼저)

    files = []
    for r in REGIONS:
        for p in sorted(by_region[r], key=_key):
            files.append((r, p))
    return files


def read_rows(p):
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    headers = [str(h).strip() if h else "" for h in rows[0]]
    return headers, rows[1:]


def dedup_key(d):
    name = re.sub(r"[\s\-_·•]", "", str(d.get("업체명") or "").strip()).lower()
    phone = re.sub(r"\D", "", str(d.get("010번호") or ""))
    if phone:
        return f"p:{name}:{phone}"
    addr = str(d.get("수집주소") or d.get("주소") or "").strip()
    return f"a:{name}:{addr[:15]}" if addr else f"n:{name}"


def main():
    files = list_files()
    print(f"파일 {len(files)}개:")
    for r, p in files:
        print(f"  [{r}] {p.name}")

    seen = set()
    merged = []
    region_count = {r: 0 for r in REGIONS}

    # Pass 1: union of all headers — 가장 풍부한 column set 확보
    union_headers = []
    for region, p in files:
        h, _ = read_rows(p)
        for col in h:
            if col and col not in union_headers:
                union_headers.append(col)
    base_headers = union_headers
    print(f"\nunion header ({len(base_headers)} cols): {base_headers}\n")

    for region, p in files:
        h, rows = read_rows(p)
        before = len(merged)
        for row in rows:
            if not row or all(c is None for c in row):
                continue
            d = {hh: v for hh, v in zip(h, row)}
            d["_지역"] = region
            k = dedup_key(d)
            if k in seen:
                continue
            seen.add(k)
            merged.append(d)
            region_count[region] += 1
        print(f"  + {p.name}: {len(merged)-before}건 신규")

    out_headers = list(base_headers) + ["지역"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "북부수도권_통합"
    ws.append(out_headers)
    for d in merged:
        row = [d.get(h) for h in base_headers] + [d.get("_지역")]
        ws.append(row)
    wb.save(str(OUT))

    print(f"\n=== 통합 완료 ===")
    print(f"총 unique: {len(merged)}건")
    print(f"지역별:")
    for r in REGIONS:
        print(f"  {r}: {region_count[r]}건")
    print(f"저장: {OUT}")


if __name__ == "__main__":
    main()
