"""Google+Naver 검색에서 010 번호 직접 추출 — 6 region × keywords."""
import asyncio, re, sys, json
from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

CLEAN = re.compile(r'[^\d]')
TAG = re.compile(r'<[^>]+>')
REGIONS = ["남양주", "양주", "포천", "의정부", "동두천", "가평"]
KEYWORDS = ["미용실", "카페", "음식점", "학원", "헬스장", "필라테스",
            "네일", "피부관리", "고깃집", "치킨", "세탁소", "약국", "부동산", "노래방"]
LEAD_DB = Path(__file__).resolve().parent.parent / "lead_db"
DESKTOP = Path.home() / "Desktop"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
SKIP = {"전화", "번호", "연락", "http", "www", "스타벅스", "맥도날드", "GS25", "CU"}


def norm(raw: str) -> str:
    d = CLEAN.sub("", raw)
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    return raw


def is_010(p: str) -> bool:
    d = CLEAN.sub("", p)
    return d.startswith("010") and len(d) == 11


def load_existing() -> set:
    files = sorted(DESKTOP.glob("북부수도권_010FINAL_*.xlsx"))
    if not files:
        return set()
    wb = openpyxl.load_workbook(str(files[-1]))
    s = set()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            s.add(CLEAN.sub("", str(row[0])))
    print(f"기존 {len(s)}건: {files[-1].name}")
    return s


def parse_html(html: str, region: str, kw: str, via: str) -> list[dict]:
    seen, out = set(), []
    for pat in [r'"telephone"\s*:\s*"([^"]+)"',
                r'tel:(01[016789][0-9\-\.]{8,12})',
                r'01[016789][-\s.]?\d{3,4}[-\s.]?\d{4}']:
        for m in re.finditer(pat, html):
            raw = m.group(1) if m.lastindex else m.group(0)
            p = norm(raw)
            if not is_010(p) or p in seen: continue
            seen.add(p)
            ctx = html[max(0, m.start()-300):m.end()+200]
            name = ""
            for np in [r'"name"\s*:\s*"([^"]{2,30})"', r'"title"\s*:\s*"([^"]{2,30})"']:
                nm = re.search(np, ctx)
                if nm and not any(s in nm.group(1) for s in SKIP):
                    name = nm.group(1); break
            if not name:
                for tok in re.findall(r'[가-힣]{2,15}', TAG.sub(" ", ctx)):
                    if not any(s in tok for s in SKIP):
                        name = tok; break
            out.append({"phone": p, "name": name, "region": region, "keyword": kw, "via": via})
    return out


async def fetch(page, url: str, region: str, kw: str, via: str, delay: float) -> list[dict]:
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=22000)
        await page.wait_for_timeout(int(delay * 1000))
        return parse_html(await page.content(), region, kw, via)
    except PWTimeout:
        print(f"  [timeout] {region} {kw} {via}"); return []
    except Exception as e:
        print(f"  [err] {region} {kw}: {e}"); return []


async def main():
    existing = load_existing()
    all_r, seen, rc = [], set(existing), {r: 0 for r in REGIONS}
    total = len(REGIONS) * len(KEYWORDS)
    print(f"\n구글 검색 ({total} 쿼리)\n")
    async with async_playwright() as pw:
        br = await pw.chromium.launch(headless=True, args=["--lang=ko-KR"])
        ctx = await br.new_context(user_agent=UA, locale="ko-KR", viewport={"width": 1280, "height": 900})
        await ctx.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,ico}", lambda r: r.abort())
        pg = await ctx.new_page()
        for i, region in enumerate(REGIONS):
            for j, kw in enumerate(KEYWORDS):
                done = i * len(KEYWORDS) + j + 1
                q = f"{region} {kw} 전화번호"
                recs = await fetch(pg, f"https://www.google.com/search?q={q}&hl=ko&num=20&gl=kr",
                                   region, kw, "google", 1.2)
                await asyncio.sleep(0.7)
                recs += await fetch(pg, f"https://search.naver.com/search.naver?query={q}&where=local",
                                    region, kw, "naver_local", 0.9)
                await asyncio.sleep(0.5)
                new = 0
                for r in recs:
                    d = CLEAN.sub("", r["phone"])
                    if d not in seen:
                        seen.add(d); all_r.append(r); rc[region] += 1; new += 1
                if new: print(f"  [{done}/{total}] {region} {kw} +{new} → 누적 {len(all_r)}건")
        await br.close()
    hdrs = ["지역", "업종", "업체명", "010번호", "검색방법"]
    for out in [LEAD_DB / f"북부수도권_구글010_{TS}.xlsx", DESKTOP / f"북부수도권_구글010_{TS}.xlsx"]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(hdrs)
        for r in all_r: ws.append([r["region"], r["keyword"], r["name"], r["phone"], r["via"]])
        wb.save(str(out))
    bk = LEAD_DB / f"북부수도권_구글010_{TS}.json"
    with open(str(bk), "w", encoding="utf-8") as f: json.dump(all_r, f, ensure_ascii=False, indent=2)
    print(f"\n=== 완료 ===\n총 {len(all_r)}건")
    for reg in REGIONS: print(f"  {reg}: {rc[reg]}건")
    print(f"저장: 북부수도권_구글010_{TS}.xlsx")


if __name__ == "__main__":
    asyncio.run(main())
